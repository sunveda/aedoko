#!/usr/bin/env python3
"""Normalize the reviewed Tokyo municipal downloads into static app artifacts."""

from __future__ import annotations

import csv
import hashlib
import json
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


SITE_ROOT = Path(__file__).resolve().parents[1]
WORKSPACE_ROOT = SITE_ROOT.parent
DATA_ROOT = WORKSPACE_ROOT / "data" / "tokyo-aed"
OUT_ROOT = SITE_ROOT / "public" / "data"


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()


def normalized_header(value) -> str:
    return clean(value).translate(str.maketrans("（）＿ＩＤＮＯ", "()_IDNO")).lower()


def decode_text(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-16", "cp932"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Unsupported text encoding: {path}")


def header_index(rows: list[list]) -> int | None:
    scored = []
    for index, row in enumerate(rows[:20]):
        cells = [normalized_header(value) for value in row]
        hits = sum(any(token in cell for token in ("名称", "施設名", "住所", "所在地", "緯度", "経度")) for cell in cells)
        scored.append((hits * 10 + sum(bool(cell) for cell in cells), -index, index))
    if not scored or max(scored)[0] < 5:
        return None
    return max(scored)[2]


def table_rows(path: Path) -> list[dict[str, str]]:
    tables: list[list[list]] = []
    if path.suffix.lower() in {".xlsx", ".xls"} or path.read_bytes()[:4] == b"PK\x03\x04":
        workbook = load_workbook(BytesIO(path.read_bytes()), read_only=True, data_only=True)
        tables = [[list(row) for row in sheet.iter_rows(values_only=True)] for sheet in workbook.worksheets]
    else:
        tables = [list(csv.reader(decode_text(path).splitlines()))]

    output = []
    for rows in tables:
        index = header_index(rows)
        if index is None:
            continue
        headers = [normalized_header(value) for value in rows[index]]
        for row in rows[index + 1 :]:
            values = [clean(value) for value in row]
            if not any(values):
                continue
            output.append({header: values[position] if position < len(values) else "" for position, header in enumerate(headers) if header})
    return output


def json_rows(path: Path) -> list[dict]:
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(payload, dict) and isinstance(payload.get("features"), list):
        output = []
        for feature in payload["features"]:
            properties = {normalized_header(k): clean(v) for k, v in (feature.get("properties") or {}).items()}
            coordinates = (feature.get("geometry") or {}).get("coordinates") or []
            if len(coordinates) >= 2:
                properties["経度"] = coordinates[0]
                properties["緯度"] = coordinates[1]
            output.append(properties)
        return output
    if isinstance(payload, list):
        return [{normalized_header(k): clean(v) for k, v in row.items()} for row in payload if isinstance(row, dict)]
    return []


def first(row: dict, *keys: str) -> str:
    for key in keys:
        value = clean(row.get(normalized_header(key)))
        if value:
            return value
    return ""


def coordinate(value: str) -> float | None:
    match = re.search(r"-?\d+(?:\.\d+)?", clean(value).replace(",", ""))
    return float(match.group()) if match else None


def bool_value(value: str) -> bool | None:
    value = clean(value).lower()
    if value in {"有", "あり", "true", "1", "○", "可", "対応"}:
        return True
    if value in {"無", "なし", "false", "0", "×", "不可", "非対応"}:
        return False
    return None


def normalize_row(row: dict, item: dict, fetched_at: str) -> dict | None:
    latitude = coordinate(first(row, "緯度", "latitude", "lat"))
    longitude = coordinate(first(row, "経度", "longitude", "lon", "lng"))
    if latitude is None or longitude is None or not (27 <= latitude <= 36.5 and 136 <= longitude <= 142.5):
        return None

    name = first(row, "名称", "施設名", "施設名称", "設置場所", "name")
    address = first(row, "所在地_連結表記", "住所", "所在地", "address")
    if not address:
        address = "".join(first(row, key) for key in ("所在地_都道府県", "所在地_市区町村", "所在地_町字", "所在地_番地以下", "建物名等(方書)"))
    if not name:
        name = first(row, "建物名等(方書)") or "AED設置場所"
    if not address:
        return None

    placement = first(row, "設置位置", "設置場所概要(建物名・設置位置)", "方書", "その他")
    days = first(row, "利用可能曜日")
    opens = first(row, "開始時間")
    closes = first(row, "終了時間")
    availability_note = first(row, "利用可能日時特記事項", "使用可能日時", "開館時間", "備考")
    resource = item["resource"]
    source_key = f'{item["municipalityCode"]}|{round(latitude, 6)}|{round(longitude, 6)}|{name}|{address}'
    record_id = "aed-" + hashlib.sha256(source_key.encode("utf-8")).hexdigest()[:16]
    explicit_24h = "24時間" in resource.get("name", "") or "24時間" in availability_note

    return {
        "id": record_id,
        "municipalityCode": item["municipalityCode"],
        "municipalityNameJa": item["municipalityNameJa"],
        "nameJa": name,
        "nameKana": first(row, "名称_カナ", "ふりがな") or None,
        "addressJa": address,
        "latitude": latitude,
        "longitude": longitude,
        "placementJa": placement or None,
        "availableDaysRaw": days or None,
        "opensAt": opens or None,
        "closesAt": closes or None,
        "availabilityNotesJa": availability_note or None,
        "explicit24Hours": explicit_24h,
        "pediatricCapable": bool_value(first(row, "小児対応設備の有無", "小児対応機器設備の有無")),
        "externalUseAllowed": False if bool_value(first(row, "外部利用不可")) is True else None,
        "coverageType": "unknown",
        "source": {
            "sourceId": item["sourceId"],
            "datasetTitle": item["datasetTitle"],
            "datasetUrl": item["datasetUrl"],
            "resourceUrl": resource["url"],
            "publisher": item["publisher"],
            "license": item.get("licenseTitle") or item.get("licenseId") or "See source",
            "sourceUpdatedAt": resource.get("lastModified") or item.get("metadataModified"),
            "fetchedAt": fetched_at,
        },
    }


def main() -> None:
    manifest = json.loads((DATA_ROOT / "download-manifest.json").read_text(encoding="utf-8"))
    municipalities = json.loads((DATA_ROOT / "tokyo-municipalities.json").read_text(encoding="utf-8"))
    fetched_at = manifest["generatedAt"]
    records_by_id: dict[str, dict] = {}
    rejected = defaultdict(int)
    included_sources: dict[str, dict] = {}

    for item in manifest["downloads"]:
        if item["status"] != "downloaded":
            continue
        path = WORKSPACE_ROOT / item["localPath"]
        try:
            if path.suffix.lower() in {".json", ".geojson"}:
                rows = json_rows(path)
            elif path.suffix.lower() in {".csv", ".xlsx", ".xls"} or path.read_bytes()[:4] in {b"PK\x03\x04", b"\xd0\xcf\x11\xe0"}:
                rows = table_rows(path)
            else:
                continue
        except Exception as error:
            print(f"skip parse error: {path}: {error}", file=sys.stderr)
            continue

        accepted = 0
        for row in rows:
            record = normalize_row(row, item, fetched_at)
            if record is None:
                rejected[item["municipalityCode"]] += 1
                continue
            existing = records_by_id.get(record["id"])
            if existing and record["explicit24Hours"]:
                existing["explicit24Hours"] = True
                existing["availabilityNotesJa"] = existing.get("availabilityNotesJa") or "24時間使用可能リスト掲載"
            elif not existing:
                records_by_id[record["id"]] = record
            accepted += 1
        if accepted:
            included_sources[item["sourceId"]] = {
                "sourceId": item["sourceId"],
                "municipalityCode": item["municipalityCode"],
                "municipalityNameJa": item["municipalityNameJa"],
                "datasetTitle": item["datasetTitle"],
                "datasetUrl": item["datasetUrl"],
                "resourceUrl": item["resource"]["url"],
                "publisher": item["publisher"],
                "license": item.get("licenseTitle") or item.get("licenseId") or "See source",
            }

    records = sorted(records_by_id.values(), key=lambda row: (row["municipalityCode"], row["nameJa"], row["id"]))
    count_by_municipality = defaultdict(int)
    source_ids_by_municipality = defaultdict(set)
    for record in records:
        count_by_municipality[record["municipalityCode"]] += 1
        source_ids_by_municipality[record["municipalityCode"]].add(record["source"]["sourceId"])

    discovered_codes = {item["municipalityCode"] for item in manifest["downloads"]}
    coverage = []
    for municipality in municipalities:
        code = municipality["code"]
        count = count_by_municipality[code]
        status = "validated" if count else ("partial" if code in discovered_codes else "missing")
        note = None
        if status == "partial":
            note = "An official source was downloaded, but no coordinate-valid records were emitted."
        elif status == "missing":
            note = "No reviewed reusable municipal AED source is included in this pilot snapshot."
        coverage.append({
            "municipalityCode": code,
            "nameJa": municipality["nameJa"],
            "status": status,
            "recordCount": count,
            "checkedAt": fetched_at,
            "sourceIds": sorted(source_ids_by_municipality[code]),
            "note": note,
        })

    generated_at = datetime.now(timezone.utc).isoformat()
    OUT_ROOT.mkdir(parents=True, exist_ok=True)
    (OUT_ROOT / "aed-tokyo.v1.json").write_text(json.dumps({"version": 1, "generatedAt": generated_at, "snapshotAt": fetched_at, "recordCount": len(records), "records": records}, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    (OUT_ROOT / "aed-tokyo-coverage.v1.json").write_text(json.dumps({"version": 1, "generatedAt": generated_at, "municipalityCount": 62, "sourcedMunicipalityCount": sum(row["status"] != "missing" for row in coverage), "validatedMunicipalityCount": sum(row["status"] == "validated" for row in coverage), "municipalities": coverage}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (OUT_ROOT / "aed-tokyo-attribution.v1.json").write_text(json.dumps({"version": 1, "generatedAt": generated_at, "sources": sorted(included_sources.values(), key=lambda row: (row["municipalityCode"], row["sourceId"]))}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (OUT_ROOT / "aed-tokyo-qa.v1.json").write_text(json.dumps({"version": 1, "generatedAt": generated_at, "emittedRecordCount": len(records), "rejectedRowsByMunicipality": dict(sorted(rejected.items())), "warnings": ["Availability defaults to unknown unless the source explicitly identifies 24-hour access.", "Coordinates point to source-listed locations and do not verify a physical entrance."]}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"records": len(records), "validatedMunicipalities": sum(row["status"] == "validated" for row in coverage), "sources": len(included_sources), "rejectedRows": sum(rejected.values())}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
